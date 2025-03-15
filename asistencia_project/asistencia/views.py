from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth import authenticate, login as auth_login, logout 
from django.views.decorators.csrf import csrf_exempt
from .models import *
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from .forms import AsistenciaForm
from django.contrib.auth.decorators import login_required, user_passes_test
from .forms import *
import datetime
import json
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill


# Create your views here.
@csrf_exempt
def acceso(request):  # Cambia el nombre de la vista a "acceso"
    context = {}
    
    if request.method == 'POST':
        # Verifica que los campos del formulario estén presentes
        if 'inputUsuario' in request.POST and 'inputContrasena' in request.POST:
            usuario = request.POST.get('inputUsuario')
            clave = request.POST.get('inputContrasena')
            
            # Autentica al usuario
            user = authenticate(request, username=usuario, password=clave)
            
            if user is not None:
                # Si las credenciales son válidas, inicia sesión
                auth_login(request, user)  # Usa auth_login en lugar de login
                return redirect('/inicio')
            else:
                # Si las credenciales son inválidas, muestra un mensaje de error
                context['error'] = 'Credenciales inválidas. Intente nuevamente.'
    
    # Renderiza la plantilla de login (tanto para GET como para POST con errores)
    return render(request, 'login.html', context)





# Función auxiliar para obtener el color según el porcentaje de asistencia
def get_color_porcentaje(porcentaje):
    if porcentaje < 50:
        return 'red'
    elif 50 <= porcentaje <= 75:
        return 'yellow'
    else:
        return 'green'

@login_required
def inicio(request):
    usuario = request.user

    # Obtener el rango de tiempo (marzo a diciembre del año actual)
    año_actual = timezone.now().year
    fecha_inicio = timezone.datetime(año_actual, 3, 1).date()  # 1 de marzo
    fecha_fin = timezone.datetime(año_actual, 12, 31).date()  # 31 de diciembre

    # Obtener todos los estudiantes del curso del profesor (si es profesor)
    if usuario.tipo_usuario == 'profesor' and usuario.curso:
        estudiantes = CustomUser.objects.filter(tipo_usuario='estudiante', curso=usuario.curso)
    elif usuario.tipo_usuario == 'estudiante':
        estudiantes = CustomUser.objects.filter(tipo_usuario='estudiante', id=usuario.id)
    else:
        estudiantes = CustomUser.objects.none()

    # Obtener todos los domingos en el rango de fechas
    domingos = []
    current_date = fecha_inicio
    while current_date <= fecha_fin:
        if current_date.weekday() == 6:  # Domingo es 6 en Python
            domingos.append(current_date)
        current_date += datetime.timedelta(days=1)

    # Calcular el porcentaje de asistencia para cada domingo
    eventos = []
    for domingo in domingos:
        asistencias = Asistencia.objects.filter(fecha=domingo, estudiante__in=estudiantes)
        total_estudiantes = estudiantes.count()
        total_presentes = asistencias.filter(estado='presente').count()
        porcentaje = (total_presentes / total_estudiantes) * 100 if total_estudiantes > 0 else 0

        eventos.append({
            'title': f"{porcentaje:.2f}%",  # Mostrar el porcentaje
            'start': domingo.isoformat(),  # Fecha del domingo
            'color': get_color_porcentaje(porcentaje),  # Color según el porcentaje
        })

    # Convertir la lista de eventos a JSON
    eventos_json = json.dumps(eventos)

    context = {
        'usuario': usuario,
        'eventos': eventos_json,  # Pasar los eventos como JSON
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    }

    return render(request, 'inicio.html', context)




@login_required
def modificar_asistencia(request, asistencia_id):
    asistencia = get_object_or_404(Asistencia, id=asistencia_id)

    if request.method == 'POST':
        form = AsistenciaForm(request.POST, instance=asistencia)
        if form.is_valid():
            form.save()
            return redirect('inicio')
    else:
        form = AsistenciaForm(instance=asistencia)

    context = {
        'form': form,
        'asistencia': asistencia,
    }

    return render(request, 'modificar_asistencia.html', context)


@login_required
@login_required
def modificar_asistencia_fecha(request, fecha):
    # Convertir la fecha de string a objeto date
    fecha_obj = datetime.datetime.strptime(fecha, '%Y-%m-%d').date()

    # Obtener todos los estudiantes del curso del profesor (si es profesor)
    if request.user.tipo_usuario == 'profesor' and request.user.curso:
        estudiantes = CustomUser.objects.filter(tipo_usuario='estudiante', curso=request.user.curso)
    else:
        estudiantes = CustomUser.objects.filter(tipo_usuario='estudiante', id=request.user.id)

    # Obtener o crear asistencias para la fecha seleccionada
    asistencias = []
    for estudiante in estudiantes:
        asistencia, created = Asistencia.objects.get_or_create(
            estudiante=estudiante,
            fecha=fecha_obj,
            defaults={'estado': 'presente'}  # Estado predeterminado
        )
        asistencias.append(asistencia)

    if request.method == 'POST':
        # Procesar el formulario para cada asistencia
        for asistencia in asistencias:
            form = AsistenciaForm(request.POST, prefix=str(asistencia.id), instance=asistencia)
            if form.is_valid():
                form.save()
        return redirect('inicio')

    # Crear un formulario para cada asistencia
    forms = [AsistenciaForm(prefix=str(asistencia.id), instance=asistencia) for asistencia in asistencias]

    context = {
        'fecha': fecha_obj,
        'forms': forms,
    }

    return render(request, 'modificar_asistencia_fecha.html', context)


# Función para verificar si el usuario es admin
def es_admin(user):
    return user.tipo_usuario == 'admin'


@login_required
@user_passes_test(es_admin, login_url='/acceso')  # Solo usuarios admin pueden acceder
def admin_panel(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('admin_panel')  # Redirige al panel de admin después de crear el usuario
    else:
        form = CustomUserCreationForm()

    # Obtener todos los usuarios para mostrarlos en el panel
    usuarios = CustomUser.objects.all()
    context = {
        'form': form,
        'usuarios': usuarios,
    }
    return render(request, 'admin_panel.html', context)





def exportar_asistencias_excel(request):
    # Crear un libro de Excel y una hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencias"

    # Definir los encabezados
    headers = [
        "Clase", "Estudiante", "Profesor", "Administrador",
        "Días Asistidos", "Días No Asistidos", "Porcentaje de Asistencia"
    ]
    ws.append(headers)

    # Definir colores para resaltar celdas
    fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Obtener el rango de tiempo (marzo a diciembre del año actual)
    año_actual = timezone.now().year
    fecha_inicio = timezone.datetime(año_actual, 3, 1).date()  # 1 de marzo
    fecha_fin = timezone.datetime(año_actual, 12, 31).date()  # 31 de diciembre

    # Obtener todos los cursos
    cursos = Curso.objects.all()

    for curso in cursos:
        # Obtener los estudiantes del curso
        estudiantes = CustomUser.objects.filter(tipo_usuario='estudiante', curso=curso)

        for estudiante in estudiantes:
            # Obtener las asistencias del estudiante en el rango de fechas
            asistencias = Asistencia.objects.filter(estudiante=estudiante, fecha__range=(fecha_inicio, fecha_fin))
            total_asistencias = asistencias.count()
            total_domingos = ((fecha_fin - fecha_inicio).days // 7) + 1  # Total de domingos en el rango
            dias_no_asistidos = total_domingos - total_asistencias
            porcentaje_asistencia = (total_asistencias / total_domingos) * 100 if total_domingos > 0 else 0

            # Obtener el profesor y el administrador del curso
            profesor = CustomUser.objects.filter(tipo_usuario='profesor', curso=curso).first()
            admin = CustomUser.objects.filter(tipo_usuario='admin').first()

            # Agregar una fila con los datos
            row = [
                curso.nombre,
                f"{estudiante.nombre} {estudiante.apellido}",
                f"{profesor.nombre} {profesor.apellido}" if profesor else "N/A",
                f"{admin.nombre} {admin.apellido}" if admin else "N/A",
                total_asistencias,
                dias_no_asistidos,
                f"{porcentaje_asistencia:.2f}%"
            ]
            ws.append(row)

            # Resaltar la celda del porcentaje según el rango
            porcentaje_cell = ws.cell(row=ws.max_row, column=7)
            if porcentaje_asistencia < 50:
                porcentaje_cell.fill = fill_red
            elif 50 <= porcentaje_asistencia <= 75:
                porcentaje_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Amarillo
            else:
                porcentaje_cell.fill = fill_green

    # Configurar la respuesta HTTP para descargar el archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=asistencias.xlsx'
    wb.save(response)

    return response